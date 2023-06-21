import datetime
import openpyxl
from tinkoffOperation import tinkoffOperation

def get_fin_operations_list(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    rows = sheet.max_row
    fin_operations = list()
    for i in range(3, rows + 1):
        operation_type = str(sheet.cell(row=i, column=3).value)
        category = str(sheet.cell(row=i, column=4).value)
        d = str(sheet.cell(row=i, column=8).value)
        payment = float(str(sheet.cell(row=i, column=12).value))
        comment = str(sheet.cell(row=i, column=20).value)
        dt = datetime.datetime.strptime(d, '%Y-%m-%d %H:%M:%S')
        date = dt.date()
        fin_operations.append(tinkoffOperation(date, operation_type, category, payment, comment))
    return fin_operations

def write_to_excel(lst):
    wb = openpyxl.Workbook()
    wb.create_sheet(title='Тинькофф', index=0)
    sheet = wb['Тинькофф']
    sheet.append(['Сальдо на начало периода'])
    sheet.append(['Дата', 'Поступление', 'Внесение л.с', 'Перевод/снятие л.с', 'Налоги/взносы', 'Расходы', 'Назначение платежа','Сальдо на конец периода'])
    row = 3
    for operation in lst:
        value = operation.date
        cell = sheet.cell(row=row, column=1)
        cell.value = value
        pm = operation.payment
        com = operation.comment
        cell = sheet.cell(row=row, column=7)
        cell.value = com
        if operation.operationType == 'Credit':
            cell = sheet.cell(row=row, column=2)
            cell.value = pm
        else:
            if operation.category == 'contragentPeople':
                cell = sheet.cell(row=row, column=4)
                cell.value = pm
            else:
                cell = sheet.cell(row=row, column=6)
                cell.value = pm
        row = row + 1

        wb.save('example.xlsx')

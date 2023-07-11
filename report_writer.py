import openpyxl

dict_columns_names = {
    'Дата': 1,
    'Поступление': 2,
    'Доход УСН': 3,
    'Внесение л.с': 4,
    'Перевод/снятие л.с': 5,
    'Налоги/взносы': 6,
    'Расходы': 7,
    'Назначение платежа': 8
}


def write_to_excel(lst, name):
    wb = openpyxl.Workbook()
    wb.create_sheet(title='List', index=0)
    sheet = wb['List']
    sheet.append(['Сальдо на начало периода'])
    sheet.append(
        ['Дата', 'Поступление', 'Доход УСН', 'Внесение л.с', 'Перевод/снятие л.с', 'Налоги/взносы', 'Расходы',
         'Назначение платежа',
         'Сальдо на конец периода'])
    row = 3
    for operation in lst:
        value = operation.date
        cell = sheet.cell(row=row, column=dict_columns_names['Дата'])
        cell.value = value
        pm = operation.payment
        com = operation.comment
        counterparty = operation.counterparty
        cell = sheet.cell(row=row, column=dict_columns_names['Назначение платежа'])
        cell.value = com
        if operation.operationType == 'Credit':
            if 'эквайринг' in com.lower() or 'МБК' in com:
                cell = sheet.cell(row=row, column=dict_columns_names['Поступление'])
                cell.value = pm
            elif name in counterparty.lower() or name in com.lower():
                cell = sheet.cell(row=row, column=dict_columns_names['Внесение л.с'])
                cell.value = pm
            else:
                cell = sheet.cell(row=row, column=dict_columns_names['Доход УСН'])
                cell.value = pm
        else:
            if name in counterparty.lower() or 'снятие по карте' in com.lower():
                cell = sheet.cell(row=row, column=dict_columns_names['Перевод/снятие л.с'])
                cell.value = pm
            elif 'УФК' in counterparty or 'ФНС' in counterparty:
                cell = sheet.cell(row=row, column=dict_columns_names['Налоги/взносы'])
                cell.value = pm
            else:
                cell = sheet.cell(row=row, column=dict_columns_names['Расходы'])
                cell.value = pm
        row = row + 1

        wb.save('example.xlsx')


def test(lst):
    for i in lst:
        if i.counterparty == '':
            print('empty')

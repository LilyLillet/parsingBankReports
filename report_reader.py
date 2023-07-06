import datetime
import re

import openpyxl
import warnings

import pandas

from finOperation import finOperation

warnings.simplefilter("ignore")

dict_tinkoff = {
    'operation_type': 3,
    'counterparty': 13,
    'date': 8,
    'payment': 12,
    'comment': 20
}


def open_excel_workbook(file_path):
    if 'tinkoff' in file_path:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        return sheet
    elif 'alfa' in file_path:
        gg = pandas.read_excel(file_path, skiprows=10)
        return gg
    elif 'sber' in file_path:
        gg = pandas.read_excel(file_path, skiprows=9)
        return gg


def get_tinkoff_operations_list(file_path):
    sheet = open_excel_workbook(file_path)
    rows = sheet.max_row
    fin_operations = list()
    for i in range(3, rows + 1):
        operation_type = str(sheet.cell(row=i, column=dict_tinkoff['operation_type']).value)
        counterparty = str(sheet.cell(row=i, column=dict_tinkoff['counterparty']).value)
        d = str(sheet.cell(row=i, column=dict_tinkoff['date']).value)
        payment = float(str(sheet.cell(row=i, column=dict_tinkoff['payment']).value))
        comment = str(sheet.cell(row=i, column=dict_tinkoff['comment']).value)
        dt = datetime.datetime.strptime(d, '%Y-%m-%d %H:%M:%S')
        date = dt.date()
        fin_operations.append(finOperation(date, operation_type, counterparty, payment, comment))
    return fin_operations


def get_alfa_operations_list(file_path):
    gg = open_excel_workbook(file_path)
    new_columns = []
    second_row = gg.iloc[0]
    for i in range(len(gg.columns)):
        if 'Unnamed' in gg.columns[i]:
            new_columns.append(second_row[i])
        else:
            new_columns.append(gg.columns[i])
    alfa_data = gg.iloc[1:]
    alfa_data.columns = new_columns
    fin_operations = list()

    for row in alfa_data.iterrows():
        dt = datetime.datetime.strptime(row[1]['Дата'], '%d.%m.%Y')
        date = dt.date()
        counterparty = row[1]['Контрагент']
        comment = row[1]['Назначение платежа']
        dbt = row[1]['Дебет']
        crdt = row[1]['Кредит']
        if is_digit(dbt):
            operation_type = 'Debit'
            payment = float(dbt)
        else:
            operation_type = 'Credit'
            payment = float(crdt)
        fin_operations.append(finOperation(date, operation_type, counterparty, payment, comment))
    return fin_operations


def get_sber_operations_list(file_path):
    gg = open_excel_workbook(file_path)
    new_columns = []
    second_row = gg.iloc[0]
    for i in range(len(gg.columns)):
        if 'Unnamed' in gg.columns[i]:
            new_columns.append(second_row[i])
        else:
            new_columns.append(gg.columns[i])
    sber_data = gg.iloc[1:]
    sber_data.columns = new_columns
    print(sber_data.columns)
    fin_operations = list()

    for row in sber_data.iterrows():
        dt = datetime.datetime.strptime(str(row[1]['Дата проводки']), '%Y-%m-%d %H:%M:%S')
        date = dt.date()
        print(date)
        comment = row[1]['Назначение платежа']
        print(comment)
        dbt = row[1]['Сумма по дебету']
        print(dbt)
        crdt = row[1]['Сумма по кредиту']
        print(crdt)
        if is_digit(dbt):
            operation_type = 'Debit'
            counterparty = row[1]['Счет']
            print(counterparty)
            payment = float(dbt)
        else:
            operation_type = 'Credit'
            counterparty = row[1]['Кредит']
            print(counterparty)
            payment = float(crdt)
        fin_operations.append(finOperation(date, operation_type, counterparty, payment, comment))
    return fin_operations


def is_digit(str):
    try:
        int(str) or float(str)
    except ValueError:
        return False
    return True
